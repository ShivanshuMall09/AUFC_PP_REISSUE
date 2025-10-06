# -*- coding: utf-8 -*-
"""
Created on Tue Sep 27 19:58:47 2022

@author: rajeev.jadaun
"""


"""
Created on Thu Nov  5 16:22:49 2020

@author: rajeev.jadaun
    AUF_CREDIT_PP_REISSUE_V.1.3.24.py:17-09-2025[Handling the problem in DL awb allocation]                                                                   
    AUF_CREDIT_PP_REISSUE_V.1.3.23.py:28-02-2025[changing the version of batccard]
    AUF_CREDIT_PP_REISSUE_V.1.3.22.py:28-01-2025[removing chip name column from batchcard]
    AUF_CREDIT_PP_REISSUE_V.1.3.20.py:28-01-2025[changing the product varient list file name to reissue product varient list]
    AUF_CREDIT_PP_REISSUE_V.1.3.17.py:23-01-2025[storing copal batchcard outside folder]
    AUF_CREDIT_PP_REISSUE_V.1.3.16.py:07-12-2024[bank Task no 68]
    AUF_CREDIT_PP_REISSUE_V.1.3.15.py:17-08-2024[Creating the batchcard file for the copal]
    AUF_CREDIT_PP_REISSUE_V.1.3.14.py:05-07-2024[changing the ff format]
    AUF_CREDIT_PP_REISSUE_V.1.3.13.py:30-07-2024[changes in batchcard and the making file read only ff and mis both ]
    AUF_CREDIT_PP_REISSUE_V.1.3.12.py:24-07-2024[NEW REFERENCE NUM LOGOC CURRENT (Courier code + AUCPP + Left 11 digit of AWB NO) NEW :(Courier code + AUCPP + Right 11 digit of AWB NO])
    AUF_CREDIT_PP_REISSUE_V.1.3.11.py:17-07-2024[Changing the RTO addres in Delhivery courier connection file request it raised by aditya on mail and also adding the new issue in the file name]
    AUF_CREDIT_PP_REISSUE_V.1.3.10.py:22-04-2024[Fixxing the problem in expiry date in PP_file in config]
    AUF_CREDIT_PP_REISSUE_V.1.3.9.py:19-04-2024[fixxing the problem in the refrence no]
    AUF_CREDIT_PP_REISSUE_V.1.3.6.py:30-03-2024[1. trnsfering file automatically to the out folder
                                                2. For delvivery awb is start from 159]
    AUF_CREDIT_PP_REISSUE_V.1.3.5.py:21-03-2024[Adding file transfer to sdrive]
    AUF_CREDIT_PP_REISSUE_V.1.3.4.py:21-03-2024[Adding chip name module in batchcard]
    AUF_CREDIT_PP_REISSUE_V.1.3.3.py:21-02-2024[Making Mis Read Only]
    AUF_CREDIT_PP_REISSUE_V.1.3.2.py:23-01-2024[Deleting the used pp value of ixigo]
    AUF_CREDIT_PP_REISSUE_V.1.3.1.py:12-01-2024[Adding the pp For Ixigo on bin 406977]
    AUF_CREDIT_PP_REISSUE_V.1.3.0.py:01-03-2023[Enhancement as per datateam]
    AUF_CREDIT_PP_REISSUE_V.1.2.8.py:23-11-2022[Courier connection report]
    AUF_CREDIT_PP_REISSUE_V.1.2.7.py:18-11-2022[IP,dl_awb assigning issue]
    AUF_CREDIT_PP_REISSUE_V.1.2.6.py:01-11-2022[AWB laert, exe stop ion 0 awb, exiry chnage in embo file]
    AUF_CREDIT_PP_REISSUE_V.1.2.5.py:12-10-2022[Serial number correction in file names]
    AUF_CREDIT_PP_REISSUE_V.1.2.4.py:12-10-2022[Masked pp number]
    AUF_CREDIT_PP_REISSUE_V.1.2.3.py:12-10-2022[handled Pincode issue in file comming differently]
    AUF_CREDIT_PP_REISSUE_V.1.2.1.py:27-09-2022[Excel files created and batch cards]
    AUF_CREDIT_PP_REISSUE_V.1.1.1.py:27-09-2022[PP Embo Sorting]
"""

from openpyxl.utils import get_column_letter
import os
import glob
import datetime
import traceback 
import pandas as pd
import sys
from pathlib import Path
from openpyxl import load_workbook

try:
    print("DATA PROCESSING STARTS..... Version-1.3.24")
    cwd = os.getcwd()
    #del_file1 = glob.glob("*.txt*")
    cwd = os.getcwd()
    
    header = 0
    
    read_only_flag = ''
    ts = datetime.datetime.now()
    ptime = ts.strftime("%d.%m.%Y_%H%M%S")
    print(ptime)
    x = (ts)
    
    with open('config/AUC_batch_series', 'r') as fin:
        batch_series_data = fin.read().splitlines(True)
        batch_series = batch_series_data[0:1]
    with open('config/AUC_batch_series', 'w') as fout:
        fout.writelines(batch_series_data[1:])
    listToStr = ' '.join([str(elem) for elem in batch_series])
    batch_number = listToStr[0:5]
    fout.close()
    
    with open('config/BD_AWB.txt', 'r') as bdfin, open('config/BD_AWB_USING.txt', 'w') as bdfout:
        contents = bdfin.readlines()
        bdfout.writelines(contents)
    bdfin.close()
    bdfout.close()
    with open('config/DL_AWB.txt', 'r') as dtfin, open('config/DL_AWB_USING.txt', 'w') as dtfout:
        contents = dtfin.readlines()
        dtfout.writelines(contents)
    dtfin.close()
    dtfout.close()
    with open('config/IP_AWB.txt', 'r') as ipfin, open('config/IP_AWB_USING.txt', 'w') as ipfout:
        contents = ipfin.readlines()
        ipfout.writelines(contents)
    ipfin.close()
    ipfout.close()
    
    with open('config/ZENITH_PP_NUMBER.csv', 'r') as zenithin, open('config/ZENITH_PP_NUMBER_USING.csv', 'w') as zenithout:
        contents = zenithin.readlines()
        zenithout.writelines(contents)
    zenithin.close()
    zenithout.close()
    with open('config/VETTA_PP_NUMBER.csv', 'r') as vettain, open('config/VETTA_PP_NUMBER_USING.csv', 'w') as vettaout:
        contents = vettain.readlines()
        vettaout.writelines(contents)
    vettain.close()
    vettaout.close()
    
    with open('config/IXIGO_PP_NUMBER.csv', 'r') as IXIGOin, open('config/IXIGO_PP_NUMBER_USING.csv', 'w') as IXIGOout:
        contents = IXIGOin.readlines()
        IXIGOout.writelines(contents)
    IXIGOin.close()
    IXIGOout.close()
    
    # sorting defination for EMBOSSA file
    
    def convert_to_read_only(file_path):
        password = 'your_password_here' 
        global read_only_flag
        
        if read_only_flag == 'NO':
            return
        else:
            try:
                wb = load_workbook(file_path)

                # Protect all sheets with password
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    ws.protection.sheet = True
                    ws.protection.password = password
                
                # Set sheet zoom to 85%
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    ws.sheet_view.zoomScale = 85
                
                # Auto-fit all columns
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for col in range(1, ws.max_column + 1):
                        column_letter = get_column_letter(col)
                        max_length = 0
                        for cell in ws[column_letter]:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        # Set column width to accommodate the maximum content length
                        ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2
                
                # Save the modified workbook
                wb.save(file_path)
            except Exception as e:
                print(f"Error converting '{file_path}': {e}")
            
    def my_sort_embo(line):
        line = line.strip()
        couriersort = line[-2:]
        # print(couriersort)
        return couriersort
    
    
    
    def my_sort_ff(line):
        line_fields = line.strip().split('|')
        couriersort = (line_fields[10])
        return couriersort
    
    
    def courier_mode(cust_pincode):
        
        with open('config/PINCODE_MASTER.csv', 'r') as pin:
            routing_codewe = pin.readlines()
            flag = 0
            routing_code = ''
            for line in routing_codewe:
                pincodedata = line.split(',')
                if cust_pincode != '':
                    # and pincodedata[4]=='BLUEDART':
                    if cust_pincode in pincodedata[0]:
                        flag = 1
                        courier = pincodedata[2].rstrip()
                        routing_code = pincodedata[1].rstrip()
                        break
            if flag == 0:
                courier = 'Speedpost'
                routing_code = '921-417'
                courier = str(courier)
        pin.close()
        return(courier,routing_code)
    
    def awb_assign(cust_courier):
        awb_number = ''
        #print(cust_courier)
        if cust_courier.upper() == 'BLUEDART':
            #print("yesy")
            with open('config/BD_AWB_USING.txt', 'r') as bdfin:
                bdawb = bdfin.readline()
                #print(bdawb)
                awb_number=bdawb.rstrip()
                bdawb1 = bdfin.readlines()
                bdremainingnew = len(bdawb1)
                # print(bdremainingnew)
                if bdremainingnew < 1:
                    with open ("Error.txt",'w') as error_file:
                        error_file.write("!!!!BD AWB NUMBER FINISHED!!!!!!")
                    error_file
                    
                    sys.exit("!!!!BD AWB NUMBER FINISHED!!!!!!")
                    
            with open('config/BD_AWB_USING.txt', 'w') as bdfout:
                 bdfout.writelines(bdawb1)
                 
                        
        elif cust_courier.upper() == 'DELHIVERY':
            with open('config/DL_AWB_USING.txt', 'r') as dlfin:
                dlawb = dlfin.readline()
                awb_number=dlawb.rstrip()
                dlawb_1 = dlfin.readlines()
                awbremainingnew = len(dlawb_1)
                # print(awbremainingnew)
                if awbremainingnew < 1:
                    with open ("Error.txt",'w') as error_file:
                        error_file.write("!!!!DELHIVERY AWB NUMBER FINISHED!!!!!!")
                    error_file
                    sys.exit("!!!!DELHIVERY AWB NUMBER FINISHED!!!!!!")
            dlawb_2 = list(map(str, dlawb_1))
            dlawb_2 = ''.join(dlawb_2)        
            with open('config/DL_AWB_USING.txt', 'w') as dlfout:
                dlfout.write(dlawb_2)
                 
                
        elif cust_courier.upper() == 'SPEEDPOST':
            with open('config/IP_AWB_USING.txt', 'r') as ipfin:
                # print("speed")
                ipawb = ipfin.readline()
                awb_number = ipawb.rstrip()
                ipawb_1 = ipfin.readlines()
                ipawbremainingnew = len(ipawb_1)
                # print(ipawbremainingnew)
                if ipawbremainingnew < 1:
                    with open ("Error.txt",'w') as error_file:
                        error_file.write("!!!!IP AWB NUMBER FINISHED!!!!!!")
                    error_file
                    sys.exit("!!!!IP AWB NUMBER FINISHED!!!!!!")
            ipawb_2 = list(map(str, ipawb_1))
            ipawb_2 = ''.join(ipawb_2)
            
            with open('config/IP_AWB_USING.txt', 'w') as ipfout:
                ipfout.write(ipawb_2)
        # print(awb_number)
        return(awb_number)
                
                
    
    
    def pp_num(bin_check):
        pp_cardnumber_acc = ''
        if bin_check == '465523' :#Vetta cards
            with open('config/VETTA_PP_NUMBER_USING.csv', 'r') as vetta_pp_in:
                vetta_pp_total    = vetta_pp_in.readline()
                pp_cardnumber_acc = vetta_pp_total.rstrip()
                # print(pp_cardnumber_acc + 'vetta pp')
                vetta_pp_unused   = vetta_pp_in.readlines()
            vetta_pp_in.close()    
               
            with open('config/VETTA_PP_NUMBER_USING.csv', 'w') as vetta_pp_out:
                vetta_pp_out.writelines(vetta_pp_unused)
            vetta_pp_out.close()
            
        elif bin_check == '457036' or bin_check == '653023' or bin_check == '653024' :#Zenith cards
            with open('config/ZENITH_PP_NUMBER_USING.csv', 'r') as zenith_pp_in:
                zenith_pp_total   = zenith_pp_in.readline()
                pp_cardnumber_acc = zenith_pp_total.rstrip()
                # print(pp_cardnumber_acc + 'zenith pp')
                zenith_pp_unused  = zenith_pp_in.readlines()
            zenith_pp_in.close()    
            with open('config/ZENITH_PP_NUMBER_USING.csv', 'w') as zenith_pp_out:
                zenith_pp_out.writelines(zenith_pp_unused)
            zenith_pp_out.close()
        #print(pp_cardnumber_acc)
        # return pp_cardnumber_acc
    
        elif bin_check == '406977' or bin_check == '653062' :#IXIGO
            with open('config/IXIGO_PP_NUMBER_USING.csv', 'r') as IXIGO_pp_in:
                IXIGO_pp_total   = IXIGO_pp_in.readline()
                pp_cardnumber_acc = IXIGO_pp_total.rstrip()
                # print(pp_cardnumber_acc + 'zenith pp')
                IXIGO_pp_unused  = IXIGO_pp_in.readlines()
            IXIGO_pp_in.close()    
            with open('config/IXIGO_PP_NUMBER_USING.csv', 'w') as IXIGO_pp_out:
                IXIGO_pp_out.writelines(IXIGO_pp_unused)
            IXIGO_pp_out.close()
        #print(pp_cardnumber_acc)
        return pp_cardnumber_acc
        
    
    
    def pp_dup_check(number):
        pp_num = []  
        with open('config/AUF_CREDIT_PP_CONSOLE.txt', 'r') as file_read:
            contents = file_read.readlines()
            new_contents = [x.strip() for x in contents]  # Remove newline characters
            if number in new_contents:
                print("PP duplicate number found: " + number)
            elif number in pp_num:
                print("PP duplicate number found: " + number)
            else:
                # print("PP number not found in the file: " + number)
                pp_num.append(number)
                with open('config/AUF_CREDIT_PP_CONSOLE.txt','a') as file_write:
                    for line in pp_num:
                        file_write.write(line+'\n')
                file_write.close()
        return number
    
    
    def excel_convertor(name_ff):
        #ext='.txt'
        if Path(name_ff).is_file():
            with open(name_ff,'r') as ff_file ,open(name_ff[:-7]+ptime+'.txt','w') as ff_file_out:
                ff_file_out.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date|Extention|AccountHolder_Name|Product Varient|Complimentary Lounge Access|Expiry Date\n')
                                # Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date|Extention|AccountHolder_Name|Product Varient|Complimentary Lounge Access|Expiry Date\n
                
                contents  = ff_file.readlines()
                contents.sort(key=my_sort_ff)
                sr_no = 1
                for line in contents:
                    s_no = "%04d" % sr_no
                    ff_file_out.write(str(s_no)+'|'+line)
                    sr_no+=1
            ff_file.close()
            ff_file_out.close()
            sr_no-=1
            # os.remove(name_ff)
            df = pd.read_csv(name_ff[:-7]+ptime+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
            df.to_excel(name_ff[:-7]+str(sr_no).zfill(4)+'_'+str(ptime)+'.xlsx', 'Sheet1', index=False)
            os.remove(name_ff)
            os.remove(name_ff[:-7]+ptime+'.txt')
    
    def embo_sorting(name_embo):
        #ext='.txt'
        if Path(name_embo).is_file():
            with open(name_embo,'r') as embo_file_in ,open(name_embo[2:],'w') as embo_file_out:
                contents  = embo_file_in.readlines()
                # sorting using our custom logic
                contents.sort(key=my_sort_embo)
                for line in contents:
                    embo_file_out.write(line)
                    
            embo_file_in.close()
            embo_file_out.close()
            
            os.remove(name_embo)
    
    def lounge_config(logo_infile):
        
        with open('config/Lounge_config.csv', 'r') as IXIGOin:
            contents = IXIGOin.readlines()
            for line in contents:
                data = line.split(',')
                logo = data[0].rstrip()
                lounge_access = data[1].rstrip()
                if logo_infile == logo:
                    return lounge_access
                    
        
        
        return '' 
    
    def prod_var(embo_bin,embo_logo):
        with open('config/Reissue_product_varient_list.txt','r') as data_file:
            contents = data_file.readlines()
            file_embo,file_logo,file_plastic_id,file_gender,file_product,file_varient =  '','','','','',''
            embo_product =  'New Product'
            embo_varient =  'New Varient'
            for line in contents[1:]:
                line_data = line.split(',')
                file_embo = line_data[0]
                file_logo = line_data[1]
                file_product = line_data[2].upper()    
                if embo_bin == file_embo and file_logo == embo_logo:
                    
                        embo_product = file_product
                        embo_varient = file_varient
                        return(embo_product)
                        # break
                    
            
        return(embo_product)
    
    
    
    with open('config/RTO_ADDRESS.txt', 'r') as IXIGOin:
        contents = IXIGOin.readlines()
        for line in contents:
            data = line.split('|')
            rto_address = data[0].rstrip()
            rto_pin = data[1].rstrip()
            
            
    file_name = glob.glob('PP_REISSUE_*')
        
    for file in file_name:
        # print(file)
        with open(file, 'r') as filehandle:
            filecontents = filehandle.readlines()
            for line1 in filecontents[1:]:
                line = line1.split('~')
                PP_CardNumber   = line[0].rstrip()
                PP_Bin          = PP_CardNumber[0:6]
                pp_last4        = PP_CardNumber[-4:]
                CR_CardNumber   = line[1].rstrip()
                Account_Number  = line[2].rstrip()
                logo = Account_Number[3:6]
                Expiry_Date     = line[3].rstrip()
                Expiry_Date_1   = Expiry_Date[0:2]+'20'+Expiry_Date[-2:]
                Title           = line[4].rstrip()
                CardHolder_Name = line[5].rstrip()
                Customer_Name   = line[6].rstrip()
                Mailing_Address = line[7].rstrip()
                
                Address_1       = ''
                Address_2       = ''
                Address_3       = ''
                City_PIN_1      = line[8]
                try:
                    
                    City            = City_PIN_1[:-7].rstrip()
                    PIN             = City_PIN_1[-6:].rstrip()
                except:
                    print(f"Error in pincode in account number : {Account_Number}\n")
                    closeInput = input("Press ENTER to exit")
                    print ("Closing...")
                State           = line[9].rstrip() 
                Contact_Number  = line[10].rstrip() 
                Card_Action     = line[-1].rstrip()
                sep             = '|'
               
                # print(PP_Bin)
                pp_number_1       =  pp_num(PP_Bin)
                
                
                
                pp_number = pp_dup_check(pp_number_1)
                courier         =  courier_mode(PIN)[0]
                routing_code    =  courier_mode(PIN)[1]
                awb             =  awb_assign(courier)
                test            =  courier[0:2].upper()
                test            =  test.replace('SP','IP')
                courier = courier[0:2].upper()

                
                courier = courier.replace('DE', 'DL')
                courier = courier.replace('SP', 'IP')
                courier = courier.replace('BL', 'BD')
                awb_R = '' 
                if awb[0:3] == '159':
                    awb_R = awb[3:]
                    refrence_num    =  courier+'AUCPP'+ awb_R[:11]
                else:
                    refrence_num    =  courier+'AUCPP'+ awb[:11]
                
                # refrence_num    =  courier+'AUCPP'+ awb[-11:]  new logic
                
               
                product = prod_var(PP_Bin,logo)
                product = product.replace('\n', '')
                
                
                complimentory_lounge_access = lounge_config(logo)
                track1_acc = '%PP/'+Title+'/'+CardHolder_Name+'//''?'
                track2_acc = ';'+pp_number+'='+ Expiry_Date_1 +'?'
                    
                with open('xxAUC_PP_Reissue_'+str(PP_Bin)+'_L'+str(logo)+'_'+str(product)+'_'+str(ptime)+'_EM.txtt', 'a') as pp_file:
                    pp_file.write('#0#'+CR_CardNumber+'#1#'+pp_number+'#2#'+track1_acc+'#3#'+track2_acc+'#4#'+Expiry_Date+'#5#'+Title+' '+CardHolder_Name+'#6#'+courier +'\n')
                    #print (file_pp)
                pp_file.close()
                
                with open('AUC_PP_Reissue_Data_Allocation_Report_raw.txt','a') as alloctaion_file:
                    alloctaion_file.write(Account_Number+'|'+CR_CardNumber+'|'+Title+' '+Customer_Name.rstrip()+'|'+'|'+'|'+'|'+'|'+pp_number+'|'+'|'+'|'+'|'+refrence_num+'|'+awb+'|'+courier+'|'+routing_code+'|'+Card_Action+'|'+'|'+'|'+'1'+'|'+PP_Bin+'|'+logo+'|'+'|'+product+'|'+Mailing_Address+'|'+Address_2+'|'+Address_3+'|'+'|'+City+'|'+State+'|'+PIN+'|'+Contact_Number+'|'+'|'+'1\n')
                    
                with open(f'AUC-PP-Reissue-FF_{str(PP_Bin)}_L{str(logo)}__{str(product)}_raw.txt','a') as alloctaion_file:
                    
                    alloctaion_file.write(Account_Number+'|'+CR_CardNumber+'|'+Title.upper()+''+CardHolder_Name.rstrip().upper()+'|'+'|'+'|'+'|'+'|'+pp_number+'|'+'|'+'|'+'|'+refrence_num+'|'+awb+'|'+courier+'|'+routing_code+'|'+Card_Action+'|'+'|'+'|'+'1'+'|'+PP_Bin+'|'+logo+'|'+'|'+product+'|'+Mailing_Address+'|'+Address_2+'|'+Address_3+'|'+'|'+City+'|'+State+'|'+PIN+'|'+Contact_Number+'|'+'|'+'|'+'|'+'|'+Customer_Name+'|'+product+'|'+complimentory_lounge_access+'|'+Expiry_Date.strip()+'\n')
                
                with open('AUC-PP-Reissue-MIS-Format_raw.txt','a') as alloctaion_file:
                    alloctaion_file.write(Account_Number+'|'+CR_CardNumber+'|'+Title+' '+Customer_Name.rstrip()+'|'+'|'+'|'+'|'+'|'+pp_number+'|'+'|'+'|'+'|'+refrence_num+'|'+awb+'|'+courier+'|'+routing_code+'|'+Card_Action+'|'+'|'+'|'+'1'+'|'+PP_Bin+'|'+logo+'|'+'|'+product+'|'+Mailing_Address+'|'+Address_2+'|'+Address_3+'|'+'|'+City+'|'+State+'|'+PIN+'|'+Contact_Number+'|'+'|'+'|'+'|'+'\n')
                
                with open('config/Priority_Pass_Data_AUF_BANK.csv', 'a') as pp_file_auf_ha:
                    pp_file_auf_ha.write(pp_number+','+str(ptime[0:10])+','+Expiry_Date_1+','+Account_Number+','+Customer_Name.rstrip()+'\n')
                pp_file_auf_ha.close()
                
                if courier == 'BD':
                    with open('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt', 'a') as pp_file_auf_ha:
                        pp_file_auf_ha.write('|'+refrence_num+'|'+Title+' '+Customer_Name.rstrip()+'|'+Mailing_Address+'|||'+City+'|'+State+'|INDIA|'+PIN+'|'+Contact_Number+'|'+awb+'\n')
                    pp_file_auf_ha.close()
                
                elif courier == 'DL':
                    with open('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt', 'a') as pp_file_auf_ha:
                        pp_file_auf_ha.write('|'+awb+'|'+refrence_num+'|'+Title+' '+Customer_Name.rstrip()+'|'+City+'|'+State+'|INDIA|'+Mailing_Address+'|'+PIN+'|'+Contact_Number+'|'+'50|Prepaid|500|Secure Deliverables|'+rto_address+'|'+rto_pin+'|AU Small Finance Bank Ltd|AU Small Finance Bank Limited AU Centre, 3rd, 5th, 6th & 7th Floor, Sunny Trade Centre, New Atish Market, Jaipur, Rajasthan  302019|True|True\n')
                    pp_file_auf_ha.close()
                
                else:
                    with open('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt', 'a') as pp_file_auf_ha:
                        pp_file_auf_ha.write('|'+awb+'|'+refrence_num+'|'+City+'|'+PIN+'|'+Mailing_Address+'|'+'|'+'|'+'|' +'|'+Contact_Number+'||'+str(awb)+'\n')
                    pp_file_auf_ha.close()
                
                with open('Daily_Priority_Pass_Reissue_Data_'+str(ptime)+'.csv', 'a') as pp_file_data:
                    if header == 0:
                        pp_file_data.write('Account_Number,Logo,Card_Number,Card_Holder_name,Card_action,Priority_Pass_Number,PP_issuance_date\n')
                    pp_file_data.write(Account_Number+','+logo  +','+CR_CardNumber+','+Title+' '+Customer_Name.rstrip()+','+Card_Action+','+pp_number+','+str(ptime[0:10])+'\n')
                    header+=1
                pp_file_data.close()
                
            
            
        filehandle.close()
        
        
        
    IndiaPost_Courier = 'xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt'
    if Path(IndiaPost_Courier).is_file():
        records_list = []
        s_no = 0
        with open('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt', "r") as ff, open('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt', "w") as foutfile:
            foutfile.write('SNO|Barcode_Value|REFERANCE NUMBER|CITY|PINCODE|NAME|ADDRESS 1|ADDRESS 2|ADDRESS 3|ADDRESSEE EMAIL|ADDRESSEE MOBILE|SENDER MOBILE|POD REQUIRED'+'\n')
            contents = ff.readlines()
            for fi in contents:
                records_list.append(fi)
                s_no = (records_list.index(fi)+1)
                s_no = "%05d" % s_no
                if not fi.strip():
                    continue
                if fi:
                    foutfile.write(str(s_no)+fi)
        ff.close()
        foutfile.close()
        df = pd.read_csv('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        df.to_excel('AUF_Credit_PP_Reissue_IndiaPost_Courier_Connection_Report_'+str(ptime)+'--'+str(s_no)+'.xlsx', 'Sheet1', index=False)
        os.remove('xxxAUF_Credit_IndiaPost_Courier_Connection_Report_.txt')
        os.remove('AUF_Credit_IndiaPost_Courier_Connection_Report_.txt')
    
    Delhivery_Courier = 'xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt'
    if Path(Delhivery_Courier).is_file():
        records_list=[]
        s_no=0
        with open('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt',"r") as ff, open('AUF_Credit_Delhivery_Courier_Connection_Report_.txt',"w") as foutfile:
            foutfile.write('Waybill|Order No/Reference No|Consignee Name|City|State|Country|Address|Pincode|Phone/Mobile|Weight|Payment Mode|Package Amount|Product to be Shipped|Return Address|Return Pin|Seller Name|Seller Address|person_specific|address_specific'+'\n')
            contents = ff.readlines()
            for fi in contents:
                records_list.append(fi)
                s_no= (records_list.index(fi)+1)
                s_no = "%05d" % s_no
                if not fi.strip():
                    continue
                if fi:
                    foutfile.write(fi)
        ff.close()
        foutfile.close()
        
    
        df = pd.read_csv('AUF_Credit_Delhivery_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        df.to_excel('AUF_Credit_PP_Reissue_Delhivery_Courier_Connection_Report_'+str(ptime)+'--'+str(s_no)+'.xlsx', 'Sheet1',index=False)
        os.remove('xxxAUF_Credit_Delhivery_Courier_Connection_Report_.txt')
        os.remove('AUF_Credit_Delhivery_Courier_Connection_Report_.txt')
    
    Bluedart_Courier = 'xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt'
    if Path(Bluedart_Courier).is_file():
        records_list = []
        s_no = 0
        with open('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt', "r") as ff, open('AUF_Credit_Bluedart_Courier_Connection_Report_.txt', "w") as foutfile:
            foutfile.write('Refrence Number|Customer Name|ADDRESS1|ADDRESS2|ADDRESS3|CITY|STATE|COUNTRY|PINCODE|CONTACT1|AWB'+'\n')
            contents = ff.readlines()
            for fi in contents:
                records_list.append(fi)
                s_no = (records_list.index(fi)+1)
                s_no = "%05d" % s_no
                if not fi.strip():
                    continue
                if fi:
                    foutfile.write(fi)
        ff.close()
        foutfile.close()
        df = pd.read_csv('AUF_Credit_Bluedart_Courier_Connection_Report_.txt',encoding= 'unicode_escape', sep='|', dtype=object)
        df.to_excel('AUF_Credit_PP_Reissue_Bluedart_Courier_Connection_Report_'+str(ptime)+'--'+str(s_no)+'.xlsx', 'Sheet1', index=False)
        os.remove('xxxAUF_Credit_Bluedart_Courier_Connection_Report_.txt')
        os.remove('AUF_Credit_Bluedart_Courier_Connection_Report_.txt')
        
    
    with open('AUC_PP_Reissue_Data_Allocation_Report_raw.txt','r') as alloctaion_file ,open('AUC_PP_Reissue_Data_Allocation_Report_'+ptime+'.txt','w') as alloctaion_file_out:
        alloctaion_file_out.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Total PP Count\n')
        contents  = alloctaion_file.readlines()
        sr_no = 1
        
        for line in contents:
            s_no = "%04d" % sr_no
            alloctaion_file_out.write(str(s_no)+'|'+line)
            sr_no+=1
    alloctaion_file.close()
    alloctaion_file_out.close()
    sr_no-=1
    os.remove('AUC_PP_Reissue_Data_Allocation_Report_raw.txt')
    df = pd.read_csv('AUC_PP_Reissue_Data_Allocation_Report_'+ptime+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
    df.to_excel('AUC_PP_Reissue_Data_Allocation_Report_'+str(sr_no)+'_'+str(ptime)+'.xlsx', 'Sheet1', index=False)
    os.remove('AUC_PP_Reissue_Data_Allocation_Report_'+ptime+'.txt') 
    
    
    ff_files = glob.glob("AUC-PP-Reissue-FF*")
    for f in ff_files:
        excel_convertor(f)
        
    
    embo_files = glob.glob("xxAUC_PP_Reissue_*")
    for f in embo_files:
        embo_sorting(f)
    
    
    
    with open('AUC-PP-Reissue-MIS-Format_raw.txt','r') as mis_file ,open('AUC-PP-Reissue-MIS-Format_'+ptime+'.txt','w') as mis_file_out:
        mis_file_out.write('Sr. No|Account Number|Primary Card number|Customer Name|Add-on Card 1|Add-on Card 2|Add-on Card 3|Add-on Card 4|PP  Card No_Primary|PP  Card No_Addon1|PP  Card No_Addon2|PP  Card No_Addon3|Ref no.|AWB. No.|Courier|Courier Code|Card Action|Primary Count|Add-on Count|Total Cards|Bin|Logo|Gender Code|Varient|Address Line 1|Address Line 2|Address Line 3|Address Line 4|City|State Code|Postal Code|Mobile Number|Cust Unique ID|Credit Limit|Statement Date |Extention\n')
        contents  = mis_file.readlines()
        sr_no = 1
        for line in contents:
            s_no = "%04d" % sr_no
            mis_file_out.write(str(s_no)+'|'+line)
            sr_no+=1
    mis_file.close()
    mis_file_out.close()
    sr_no-=1
    os.remove('AUC-PP-Reissue-MIS-Format_raw.txt')
    df = pd.read_csv('AUC-PP-Reissue-MIS-Format_'+ptime+'.txt',encoding= 'unicode_escape', sep='|', dtype=object)
    df.to_excel('AUC-PP-Reissue-MIS-Format_'+str(sr_no).zfill(4)+'_'+str(ptime)+'.xlsx', 'Sheet1', index=False)
    os.remove('AUC-PP-Reissue-MIS-Format_'+ptime+'.txt')
    
    
    file_count = '.'
    
    os.chdir(file_count)
    names={}
    for fn in glob.glob('*AUC*.txtt'):
        with open(fn) as f:
            names[fn]=sum(1 for file_count in f if file_count.strip() and not file_count.startswith('~'))       
    with open('AUF_FILE_COUNT.csv', 'w') as f:
        f.write('File Name,File Count,Non-Replace,Replace'+'\n')
        [f.write('{0},{1}\n'.format(key, value)) for key, value in names.items()] 
        
    from prettytable import PrettyTable
    ptx = PrettyTable()
    ptx.field_names = ["Bin", "Artwork No.","JB No.", "EMBOSS Filename","Qty", "Job setup", "Printing method - Front/Back"]

    for key, value in names.items():
        val = key.split('_')
        file_bin = val[3].rstrip()
        art_work = 'Not_Found'
        job_setup = 'Not_Found'
        dg_color = 'Not_Found'
        with open('config/batchcard.txt','r') as data_file, open('config/batchcard_file.csv','r') as batch_file:
            contents = data_file.readlines() 
            datas = batch_file.readlines() 
            
            for line in contents[1:]:
                line_data = line.split(',')
                file_start_index = line_data[0].strip()
                file_end_index = line_data[1].strip()
                file_embo_name = line_data[2].strip()
                file_art_work = line_data[3].strip()
                file_job_setup = line_data[4].strip()
                file_dg_color = line_data[5].strip()
                
                # print(file_start_index+':'+file_end_index)
                if key[int(file_start_index):int(file_end_index)] == file_embo_name:
                    art_work = file_art_work
                    job_setup = file_job_setup
                    dg_color = file_dg_color
                    break
                
            for data in datas[1:]:
                value_1 = data.split(',')
                bin_num = value_1[0].rstrip()
                job_setup_batch = value_1[1].rstrip()
                file_name_batch = value_1[2].rstrip()
                logo_batch = value_1[3].rstrip()
                art_work_batch = value_1[4].rstrip()
                
                if (bin_num == file_bin and job_setup_batch == job_setup and  art_work_batch == art_work or file_name_batch in key):
                    
                    # print(file_bin)
                    logo = key.split('_')[4][1:]
                    card_action_e = key.split('_')[5]
                    # print(logo)
                    printing_front = value_1[5].rstrip()
                    ribbon_front = value_1[6].rstrip()
                    printing_back = value_1[7].rstrip()
                    ribbon_back = value_1[8].rstrip()
                    
                    with open('AUF_CREDIT_BATCHCARD_FILE.csv','a') as out_file:
                                        # CARD_ACTION|QTY            |Emboss_Filename|JOB_SETUP_NAME|Artwork_NO|F-PRINTING_Method|F-Ribbon_Color|B-PRINTING_Method|B-Ribbon_Color|BIN|SUB_BIN/LOGO
                        out_file.write('Reissue'+'|'+str(value)+'|'+key+'|'+job_setup_batch+'|'+art_work_batch+'|'+printing_front+'|'+ribbon_front+'|'+printing_back+'|'+ribbon_back+'|'+bin_num+'|'+logo+'\n')
                    break   
        
        
        ptx.add_row([file_bin, art_work,"", key,value, job_setup, dg_color])
    ptx.align = "c"
    ptd=ptx.get_string()
    
    ptx1 = PrettyTable()
    ptx1.field_names = ["Emboss Filename","Supervisor Name","Signature", "  Date  ","  Time  ","   Remark(if any)   "]
    for key, value in names.items():
     
        ptx1.add_row([key,"","","","",""])
    ptx1.align = "c"
    ptd1=ptx1.get_string()


    datedmy = str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]
    with open('AUF_CREDIT_BATCHCARD_'+str(ptime)+'.dat', 'w') as file:
      file.write('                                                '+'BANKING PERSONALISATION BATCH CARD'+'                              '+'Date: '+ptime[0:10]+'\n')
      file.write('                                                       AUC-'+ptime[0:10]+'-'+batch_number+'                                    '+'AUF CREDIT PROJECT(DI)\n\n')
      file.write(str(ptd))
      file.write('\n')
      file.write('TOTAL BATCH QUANTITY:'+str(sum(names.values()))+'\n')
      file.write('                                                   Data upload on Machine\n')
      file.write(str(ptd1))
      file.write('\n\n')
      file.write('PRP: 20.1                                              Rev No: 3.3                                            Date: 01-Feb-25\n')
      file.write('SEC-3: INTERNAL                                        Owner: Quality Control                                 Status: Issued\n\n')
      file.write('                                                       Page: 1 of 1')
      


    from fpdf import FPDF
    class PDF(FPDF):
        def header(self):
            self.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')
            self.ln(8)
    pdf = PDF()
    pdf.add_page('L')
    pdf.set_font("Courier", size=6)
    f = open('AUF_CREDIT_BATCHCARD_'+str(ptime)+'.dat', "r")

    for x in f:
        pdf.cell(50, 5, txt = x, ln=True, align = 'L')
        #pdf.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 20, type = '', link = '')
        #pdf.cell(ln=10)

    pdf.output('AUF_CREDIT_BATCHCARD_'+ptime[0:10]+'_'+batch_number+'.pdf')
    f.close()
    
    
    with open('AUF_CREDIT_BATCHCARD_FILE.csv','r') as in_file, open('AUF_CREDIT_BATCHCARD_'+ptime[0:10]+'_'+str(sum(names.values()))+'.TXT','a') as output_file:
        contents = in_file.readlines()
        output_file.write('CARD_ACTION|QTY|Emboss_Filename|JOB_SETUP_NAME|Artwork_NO|F-PRINTING_Method|F-Ribbon_Color|B-PRINTING_Method|B-Ribbon_Color|BIN|SUB_BIN/LOGO\n')
        sn = 0
        for line in contents:
            sn+=1
            output_file.write(str(sn)+'|'+line)
            
    df = pd.read_csv('AUF_CREDIT_BATCHCARD_'+ptime[0:10]+'_'+str(sum(names.values()))+'.TXT',encoding= 'unicode_escape',sep='|', dtype=object)
    df.to_excel('AUFC_BATCHCARDS_'+ptime[0:10]+'_'+batch_number+'_'+str(sum(names.values()))+'.xlsx', 'Sheet1', index=False)

    #---------------------------------DELETION LOG-----15.2.2----------------------
    from prettytable import PrettyTable
    ptx = PrettyTable()
    ptx.field_names = (["Client Name", "File Name","File Deletion Date", "Dispatch Date","Data Admin. Name", "Data Admin. Sign.","IT Person Name", "IT Person Sign."])




    for key,value in names.items():
     
        ptx.add_row(["AUF Bank", key, "", "", "Data Team", "", "IT Team", ""'\n'])
    ptx.align = "c"
    ptd=ptx.get_string()

    with open('AUF_DELETION_LOG_'+datedmy+'.dat', 'w') as file:
      file.write('\n\n                                                                        SEC-IS   |   02.02\n')
      file.write('                                                                        Rev.No.  |   3.0                                                   '+'AUF BANK\n')
      file.write('                                                                        Date :   |   15-Jul-18\n')
      file.write('                                                                 '+'( ***** DATA DELETION LOG ***** )''\n\n\n\n')
      file.write('|  Data Receiving Date - '+datedmy+'  |  '+'  Data Servier - DPP Server/MX Machine  |   '+'  Batch Qty. - '+str(sum(names.values()))+'   |  '+'  Batch No. - AUF-'+str(x)[8:10]+'-'+str(x)[5:7]+'-'+str(x)[0:4]+'--'+batch_number+'  |  '+'Status -             '+'|''\n\n\n')

      file.write(str(ptd))
      file.write('\n\n\n\n')
      #file.write('BATCH NO. OF PM TOOL FOR THIS BATCHCARD :_______________\n')

    class PDF(FPDF):
        def header(self):
            self.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 10, type = '', link = '')
            self.ln(8)

    pdf = PDF()
    pdf.add_page('L')
    pdf.set_font("Courier", size=7)
    f = open('AUF_DELETION_LOG_'+datedmy+'.dat', "r")

    for x in f:
        pdf.cell(50, 5, txt = x, ln=True, align = 'L')
        #pdf.image('config/colorplast_logo.png', x = 10, y = 5, w = 50, h = 20, type = '', link = '')
        #pdf.cell(ln=10)

    pdf.output('AUF_DELETION_LOG_'+datedmy+'.pdf')
    f.close()

    #---------------------------------DELETION LOG ENDS ---------------------------

    print("Batch Generated..........")

   
    with open('config/BD_AWB_USING.txt', 'r') as bdfin,open('config/BD_AWB.txt', 'w') as bdfout:
        contents= bdfin.readlines()
        bdfout.writelines(contents)
        bdremainingnew = len(contents)
    bdfin.close()
    bdfout.close()
    with open('config/DL_AWB_USING.txt', 'r') as dtfin,open('config/DL_AWB.txt', 'w') as dtfout:
        contents= dtfin.readlines()
        dtfout.writelines(contents)
        dtremainingnew = len(contents)    
    dtfin.close()
    dtfout.close()
    with open('config/IP_AWB_USING.txt', 'r') as ipfin,open('config/IP_AWB.txt', 'w') as ipfout:
        contents= ipfin.readlines()
        ipfout.writelines(contents)    
        ipremainingnew = len(contents)
    ipfout.close()
    ipfout.close()
    
    with open('config/ZENITH_PP_NUMBER_USING.csv', 'r') as zenithin, open('config/ZENITH_PP_NUMBER.csv', 'w') as zenithout:
        contents = zenithin.readlines()
        zenithout.writelines(contents)
    zenithin.close()
    zenithout.close()
    with open('config/VETTA_PP_NUMBER_USING.csv', 'r') as vettain, open('config/VETTA_PP_NUMBER.csv', 'w') as vettaout:
        contents = vettain.readlines()
        vettaout.writelines(contents)
    vettain.close()
    vettaout.close()
    
    with open('config/IXIGO_PP_NUMBER_USING.csv', 'r') as IXIGOin, open('config/IXIGO_PP_NUMBER.csv', 'w') as IXIGOout:
        contents = IXIGOin.readlines()
        IXIGOout.writelines(contents)
    IXIGOin.close()
    IXIGOout.close()
    
    
    file_name = glob.glob('PP_REISSUE_*') + glob.glob('AUF_CREDIT_BATCHCARD*.txt')
    for file in file_name:
        os.remove(file)
    file_name = glob.glob('*.dat')
    for file in file_name:
        os.remove(file)
    
    with open('AWB_REMAINING_COUNT_NEW.txt','w') as awb_count_new:
        awb_count_new.write('BD   : ' + str(bdremainingnew)+'\n'+'DL : '+str(dtremainingnew)+'\nIP   : '+str(ipremainingnew) )
        print('BD       |     DL         |      IP   ')
        print(str(bdremainingnew)+'    |     '+str(dtremainingnew)+'     |       '+str(ipremainingnew) )
    awb_count_new.close()
    
    with open('config/AWB_ALERT.csv', 'r') as alert_file:
        contents = alert_file.readlines()
        for line in contents:
            line1 = line.split(':')
            if line1[0] == 'BD':
                bd_alert = line1[1]
            if line1[0] == 'DT':
                dt_alert = line1[1]
            if line1[0] == 'IP':
                ip_alert = line1[1]
    
    if bdremainingnew<=int(bd_alert):
        print("ALERT!!!!!!!!!!!!!!!!!!!!!")
        print("HIGH ALERT BD AWB NUMBER REMAINING : "+str(bdremainingnew))
    if dtremainingnew<=int(dt_alert):
        print("ALERT!!!!!!!!!!!!!!!!!!!!!")
        print("HIGH ALERT DELHIVERY AWB NUMBER REMAINING : "+str(dtremainingnew))
    if ipremainingnew<=int(ip_alert):
        print("ALERT!!!!!!!!!!!!!!!!!!!!!")
        print("HIGH ALERT IP AWB NUMBER REMAINING : "+str(ipremainingnew))
    # print(os.getcwd())
    os.chdir('config/')
    # print(os.getcwd())
    os.remove('BD_AWB_USING.txt')
    os.remove('DL_AWB_USING.txt')
    os.remove('IP_AWB_USING.txt')
    os.remove('VETTA_PP_NUMBER_USING.csv')
    os.remove('ZENITH_PP_NUMBER_USING.csv')
    os.remove('IXIGO_PP_NUMBER_USING.csv')
    os.chdir(cwd)
    
    
    input_files= glob.glob('*.txtt') + glob.glob('*.xlsx')
    for f in input_files:
        # f.replace(f'{date_1}',f'{date_1}_{batch_number}')
        new_name = f.replace(f'{ptime}', f'{ptime}_{batch_number}')
        os.rename(f, new_name)
    
    current_directory = os.getcwd()
    excel_files = glob.glob(os.path.join(current_directory, '*MIS*.xlsx')) + glob.glob(os.path.join(current_directory, '*ff*.xlsx'))
    if excel_files:
        for excel_file in excel_files:
            convert_to_read_only(excel_file)
            
            
    path = ('.')
    ext = "txtt"
    
    for f in os.listdir(path):
        fpath = os.path.join(path, f)
    
        if os.path.isfile(fpath) and fpath.endswith(ext):
            
            time = datetime.datetime.fromtimestamp(os.path.getctime(fpath)).strftime("%d-%m-%Y-%H%M%S----"+str(sum(names.values())))
            name='AUF_CREDIT_PP_REISSUE_'+time
            os.makedirs(os.path.join(path, name), exist_ok=True)
            os.replace(fpath, os.path.join(path, name, f))
   
    from distutils.dir_util import copy_tree
    import os


    os.chdir(path) 
    fromDirectory = os.getcwd()
    source_directory_name = name
    source_directory = os.path.join(fromDirectory, source_directory_name)
    target_directory = "A:/Sdrive/LIVE/AU_Credit/"+name
    copy_tree(source_directory, target_directory)      
    
    
    
    path = ('.')
    current_date = datetime.datetime.now()
    dt1=current_date.strftime("%Y-%m-%d-%H%M%S---")
    
    import shutil
    destination_folder = "C:/Config/Batchcard"
    os.makedirs(destination_folder, exist_ok=True)
    files = glob.glob('*batchcard*.dat') +glob.glob('*deletion*.dat')
    for file in files:
        destination_path = os.path.join(destination_folder, os.path.basename(file))
        shutil.move(file, destination_path)

    for f in os.listdir(path):
        fpath = os.path.join(path, f)
        
        if os.path.isfile(fpath) and (fpath.endswith('xlsx') or fpath.endswith('pdf') or fpath.endswith('dat') or fpath.endswith('txt') or fpath.endswith('csv')):
           if not f.startswith('AUFC_BATCHCARDS_'):
            # time = datetime.fromtimestamp(os.path.getctime(fpath)).strftime("%d-%m-%Y-%H%M%S---"+str(qty))
                time = datetime.datetime.fromtimestamp(os.path.getctime(fpath)).strftime(dt1+str(sum(names.values())).rjust(4,'0'))
                name='AUF_CREDIT_PP_REISSUE_FF_MIS_'+time
                os.makedirs(os.path.join(path, name), exist_ok=True)
                os.replace(fpath, os.path.join(path, name, f))
    
    
    with open('config/Output_file_location.csv','r') as file:
        content = file.readlines()
        folder_location = content[0].split(',')[1].strip()
        
    os.chdir(path) 
    fromDirectory = os.getcwd()
    aaa = fromDirectory+'/'+name
    toDirectory = folder_location 
    bbb = toDirectory+name
    copy_tree(aaa, bbb)
    
except Exception as e:    
    traceback.print_exc()

ts = datetime.datetime.now()
ptime = ts.strftime("\n\n%d.%m.%Y_%H%M%S")
print(ptime)
 
closeInput = input("Press ENTER to exit")
print ("Closing...")